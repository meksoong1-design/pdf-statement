import re
import io
from datetime import datetime
from zoneinfo import ZoneInfo

import pdfplumber
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import gspread
from google.oauth2.service_account import Credentials


# =========================================================
# CONFIG
# =========================================================

USER_PASSWORD = "+123456+"
ADMIN_PASSWORD = "181920"

st.set_page_config(
    page_title="PDF Statement to Excel",
)

SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive",
]


# =========================================================
# GOOGLE SHEET FUNCTIONS
# =========================================================

@st.cache_resource
def get_gsheet():
    """
    Connect to Google Sheet using Streamlit secrets.

    Required secrets:

    [gsheet]
    spreadsheet_id = "..."

    [gcp_service_account]
    type = "service_account"
    project_id = "..."
    private_key_id = "..."
    private_key = "-----BEGIN PRIVATE KEY-----\n...\n-----END PRIVATE KEY-----\n"
    client_email = "..."
    client_id = "..."
    auth_uri = "https://accounts.google.com/o/oauth2/auth"
    token_uri = "https://oauth2.googleapis.com/token"
    auth_provider_x509_cert_url = "https://www.googleapis.com/oauth2/v1/certs"
    client_x509_cert_url = "..."
    """
    try:
        if "gcp_service_account" not in st.secrets:
            raise KeyError("ไม่พบ [gcp_service_account] ใน Streamlit secrets")

        if "gsheet" not in st.secrets:
            raise KeyError("ไม่พบ [gsheet] ใน Streamlit secrets")

        if "spreadsheet_id" not in st.secrets["gsheet"]:
            raise KeyError("ไม่พบ spreadsheet_id ใน [gsheet]")

        service_account_info = dict(st.secrets["gcp_service_account"])

        # แก้ปัญหา private_key ที่บางครั้งถูกเก็บเป็น \\n
        if "private_key" in service_account_info:
            service_account_info["private_key"] = service_account_info["private_key"].replace("\\n", "\n")

        creds = Credentials.from_service_account_info(
            service_account_info,
            scopes=SCOPES
        )

        client = gspread.authorize(creds)
        sheet = client.open_by_key(st.secrets["gsheet"]["spreadsheet_id"])
        return sheet

    except PermissionError as e:
        st.error("Service Account ไม่มีสิทธิ์เข้าถึง Google Sheet")
        st.warning(
            "ให้เปิด Google Sheet → Share → เพิ่ม client_email จาก secrets "
            "และตั้งสิทธิ์เป็น Editor"
        )
        st.exception(e)
        st.stop()

    except Exception as e:
        st.error("เชื่อมต่อ Google Sheet ไม่สำเร็จ")
        st.warning(
            "กรุณาตรวจสอบว่าใส่ st.secrets ถูกต้อง, เปิด Google Sheets API / Google Drive API แล้ว "
            "และแชร์ Google Sheet ให้ client_email ของ Service Account เป็น Editor แล้ว"
        )
        st.exception(e)
        st.stop()


def get_ws(name, rows, cols, header):
    """
    Get worksheet by name.
    If worksheet does not exist, create it and add header.
    If worksheet exists but is empty, add header.
    """
    sheet = get_gsheet()

    try:
        ws = sheet.worksheet(name)
    except gspread.WorksheetNotFound:
        ws = sheet.add_worksheet(title=name, rows=rows, cols=cols)
        ws.append_row(header)
        return ws

    values = ws.get_all_values()

    if not values:
        ws.append_row(header)
    elif values[0] != header:
        # ไม่ลบข้อมูลเก่า แต่ปรับ header แถวแรกให้ตรง
        ws.update("A1", [header])

    return ws


def load_stats():
    """
    Load stats from worksheet.
    If no data row exists, create default row.
    """
    ws = get_ws(
        "stats",
        10,
        4,
        ["login_count", "upload_count", "download_count", "last_access"]
    )

    rows = ws.get_all_values()

    if len(rows) < 2:
        ws.append_row([0, 0, 0, ""])
        return {
            "login_count": 0,
            "upload_count": 0,
            "download_count": 0,
            "last_access": None
        }

    r = rows[1]

    while len(r) < 4:
        r.append("")

    return {
        "login_count": int(r[0]) if str(r[0]).strip().isdigit() else 0,
        "upload_count": int(r[1]) if str(r[1]).strip().isdigit() else 0,
        "download_count": int(r[2]) if str(r[2]).strip().isdigit() else 0,
        "last_access": r[3] if r[3] else None
    }


def save_stats(stats):
    """
    Save stats to row 2.
    """
    ws = get_ws(
        "stats",
        10,
        4,
        ["login_count", "upload_count", "download_count", "last_access"]
    )

    ws.update("A2:D2", [[
        stats.get("login_count", 0),
        stats.get("upload_count", 0),
        stats.get("download_count", 0),
        stats.get("last_access") or ""
    ]])


def append_log(event):
    """
    Add one log row to log worksheet using Thailand time.
    """
    now = datetime.now(ZoneInfo("Asia/Bangkok")).strftime("%d/%m/%Y %H:%M:%S")

    ws = get_ws(
        "log",
        2000,
        2,
        ["timestamp", "event"]
    )

    ws.append_row([now, event])
    return now


def log_event(event):
    """
    Log event and update last_access safely.
    last_access will also use Thailand time because it comes from append_log().
    """
    now = append_log(event)

    stats = load_stats()
    stats["last_access"] = now
    save_stats(stats)

    return now


def load_logs(limit=100):
    """
    Load latest logs.
    """
    ws = get_ws(
        "log",
        2000,
        2,
        ["timestamp", "event"]
    )

    rows = ws.get_all_values()

    if len(rows) <= 1:
        return []

    data_rows = rows[1:]
    recent = data_rows[-limit:] if len(data_rows) > limit else data_rows

    logs = []

    for r in reversed(recent):
        if len(r) >= 2:
            logs.append(f"{r[0]} — {r[1]}")

    return logs


def test_google_sheet_connection():
    """
    Test connection and return sheet title.
    """
    sheet = get_gsheet()
    return sheet.title


# =========================================================
# LOGIN
# =========================================================

pwd = st.text_input("กรอกรหัสผ่าน", type="password")


if pwd == ADMIN_PASSWORD:
    st.title("Admin Dashboard")
    st.markdown("---")

    col_test_1, col_test_2 = st.columns([1, 3])

    with col_test_1:
        if st.button("Test Google Sheet"):
            title = test_google_sheet_connection()
            st.success(f"เชื่อมต่อสำเร็จ: {title}")

    with st.spinner("กำลังโหลดข้อมูลจาก Google Sheet..."):
        stats = load_stats()
        logs = load_logs()

    col1, col2, col3 = st.columns(3)

    col1.metric("เข้าใช้งาน", stats["login_count"])
    col2.metric("อัปโหลด PDF", stats["upload_count"])
    col3.metric("ดาวน์โหลด Excel", stats["download_count"])

    st.markdown("---")
    st.info(f"เข้าใช้ล่าสุด เวลาไทย: {stats['last_access'] or 'ยังไม่มีข้อมูล'}")

    st.markdown("### Log ย้อนหลัง")

    if logs:
        for line in logs:
            st.text(line)
    else:
        st.write("ยังไม่มีข้อมูล")

    st.stop()


elif pwd == USER_PASSWORD:
    try:
        stats = load_stats()
        stats["login_count"] += 1
        save_stats(stats)
        log_event("เข้าใช้งาน")
    except Exception as e:
        st.error("บันทึกข้อมูลการเข้าใช้งานลง Google Sheet ไม่สำเร็จ")
        st.exception(e)
        st.stop()


elif pwd != "":
    st.error("รหัสผ่านไม่ถูกต้อง")
    st.stop()


else:
    st.info("กรุณากรอกรหัสผ่านเพื่อใช้งาน")
    st.stop()


# =========================================================
# EXCEL STYLE CONFIG
# =========================================================

font_normal = Font(name="Arial", size=10, color="000000")
font_negative = Font(name="Arial", size=10, color="FF0000")
font_bold = Font(name="Arial", size=10, bold=True, color="000000")
font_header = Font(name="Arial", size=10, bold=True, color="000000")

fill_total = PatternFill("solid", start_color="D9D9D9", end_color="D9D9D9")

align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)

thin_side = Side(style="thin")
double_side = Side(style="double")


def thin_border():
    return Border(
        left=thin_side,
        right=thin_side,
        top=thin_side,
        bottom=thin_side
    )


fmt_integer = '#,##0_ ;[Red]\\-#,##0\\ '
fmt_decimal = '#,##0.00_ ;[Red]\\-#,##0.00\\ '
fmt_decimal_plain = '#,##0.00'
fmt_date = 'DD/MM/YYYY'

columns = {
    "date": 1,
    "monthly_pl": 2,
    "realized": 3,
    "unrealized_end_month": 4,
    "unrealized_change": 5,
    "commission_vat": 6,
    "deposit": 7,
    "withdraw": 8,
    "net_deposit": 9,
    "trade_long": 10,
    "trade_short": 11,
    "trade_total": 12,
    "open_long": 13,
    "open_short": 14,
    "open_total": 15,
    "equity": 16,
    "check": 17,
}

DATA_START_ROW = 4

SUM_COLS = [2, 3, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15]

number_formats = {
    1: fmt_date,
    2: fmt_integer,
    3: fmt_integer,
    4: fmt_integer,
    5: fmt_integer,
    6: fmt_integer,
    7: fmt_decimal,
    8: fmt_decimal,
    9: fmt_decimal_plain,
    10: fmt_decimal,
    11: fmt_decimal,
    12: fmt_decimal,
    13: fmt_decimal,
    14: fmt_decimal,
    15: fmt_decimal,
    16: fmt_decimal,
    17: fmt_integer,
}


# =========================================================
# PDF PARSING FUNCTIONS
# =========================================================

def to_number(value):
    """
    Convert text number to float.
    Handles:
    1,234.56
    -1,234.56
    (1,234.56)
    """
    if value is None:
        return None

    text = str(value).strip()

    m = re.search(r"\(([\d,]+\.?\d*)\)", text)
    if m:
        return -float(m.group(1).replace(",", ""))

    m = re.search(r"-?[\d,]+\.?\d+", text)
    if not m:
        return None

    try:
        return float(m.group().replace(",", ""))
    except Exception:
        return None


def to_int(value):
    try:
        if value in (None, ""):
            return 0
        return int(str(value).replace(",", "").strip())
    except Exception:
        return 0


def font_for(value):
    try:
        if value is not None and float(str(value).replace(",", "")) < 0:
            return font_negative
    except Exception:
        pass

    return font_normal


def find_date(text, filename=""):
    patterns = [
        r"Trading Date\s*:\s*(\d{1,2}/\d{1,2}/\d{4})",
        r"\b(\d{1,2}[/-]\d{1,2}[/-]\d{4})\b"
    ]

    for p in patterns:
        m = re.search(p, text, re.IGNORECASE)
        if m:
            return m.group(1).strip()

    m = re.search(r"(\d{4})(\d{2})(\d{2})", filename)
    if m:
        y, mo, d = m.groups()
        return f"{d}/{mo}/{y}"

    return "Unknown"


def table_type(header_row):
    if not header_row:
        return "unknown"

    h = " ".join(str(c) for c in header_row if c).lower()

    if "trade id" in h:
        return "futures_confirmation"

    if "unreali" in h and "trade date" in h:
        return "open_position"

    return "summary"


def read_cash(tables):
    dep = 0.0
    wd = 0.0
    found = False

    for table in tables:
        for row in table:
            if not row or not row[0]:
                continue

            for line in str(row[0]).split("\n"):
                lu = line.upper()
                nums = re.findall(r"\([\d,]+\.?\d*\)|-?[\d,]+\.?\d+", line)

                if not nums:
                    continue

                amt = to_number(nums[-1])

                if amt is None:
                    continue

                if "DEPOSIT" in lu and "SEG" in lu:
                    dep += abs(amt)
                    found = True

                elif "WITHDRAWAL" in lu and "SEG" in lu:
                    wd += abs(amt)
                    found = True

    return (
        dep if found and dep > 0 else None,
        wd if found and wd > 0 else None
    )


def read_pdf(file_bytes, filename):
    d = {
        "date": None,
        "realized": None,
        "unrealized_end_month": None,
        "commission": None,
        "vat": None,
        "commission_vat": None,
        "equity": None,
        "deposit": None,
        "withdraw": None,
        "trade_long": 0,
        "trade_short": 0,
        "open_long": 0,
        "open_short": 0,
    }

    full_text = ""
    all_tables = []

    with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
        for page in pdf.pages:
            full_text += (page.extract_text() or "") + "\n"

            page_tables = page.extract_tables()
            all_tables.extend(page_tables)

            for table in page_tables:
                if not table:
                    continue

                ttype = table_type(table[0])

                for row in table:
                    if not row:
                        continue

                    c0 = str(row[0]).strip() if row[0] else ""
                    c0l = c0.lower()
                    first = c0.split("\n")[0].lower()

                    if ttype == "futures_confirmation" and c0l == "total":
                        d["trade_long"] += to_int(row[3] if len(row) > 3 else 0)
                        d["trade_short"] += to_int(row[4] if len(row) > 4 else 0)
                        continue

                    if ttype == "open_position" and c0l == "total":
                        d["open_long"] += to_int(row[2] if len(row) > 2 else 0)
                        d["open_short"] += to_int(row[3] if len(row) > 3 else 0)
                        continue

                    if first.startswith("commission") and "vat" in c0l and "f&o" not in first:
                        nums = re.findall(r"\([\d,]+\.?\d*\)|-?[\d,]+\.?\d+", c0)

                        if len(nums) >= 2:
                            d["commission"] = to_number(nums[1])

                        if len(nums) >= 4:
                            d["vat"] = to_number(nums[3])

                        if len(nums) >= 6:
                            d["realized"] = to_number(nums[5])

                    elif "f&o unrealized" in c0l:
                        nums = re.findall(r"\([\d,]+\.?\d*\)|-?[\d,]+\.?\d+", c0)

                        if len(nums) >= 2:
                            d["unrealized_end_month"] = to_number(nums[1])

                    elif c0l == "total equity":
                        for ci in [2, 1]:
                            v = row[ci] if len(row) > ci else None

                            if v and re.search(r"[\d(]", str(v)):
                                d["equity"] = to_number(v)
                                break

    fallback_patterns = [
        (
            "commission",
            r"^Commission\s+([\d,().]+)\s+([\d,().]+)"
        ),
        (
            "vat",
            r"^VAT\s+([\d,().]+)\s+([\d,().]+)"
        ),
        (
            "realized",
            r"F&O Realized P&L\s+([\d,().]+)\s+([\d,().]+)"
        ),
        (
            "unrealized_end_month",
            r"F&O Unrealized Profit/Loss\s+([\d,().]+)\s+([\d,().]+)"
        ),
        (
            "equity",
            r"Total Equity\s+([\d,]+\.?\d*)\s+([\d,]+\.?\d*)"
        ),
    ]

    for key, pattern in fallback_patterns:
        if d[key] is None:
            m = re.search(pattern, full_text, re.M)
            if m:
                d[key] = to_number(m.group(2))

    commission = d["commission"] or 0
    vat = d["vat"] or 0

    if d["commission"] is not None or d["vat"] is not None:
        d["commission_vat"] = commission + vat

    d["date"] = find_date(full_text, filename)
    d["deposit"], d["withdraw"] = read_cash(all_tables)

    return d


# =========================================================
# EXCEL GENERATION
# =========================================================

def make_excel(records):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    for r, h in {
        1: 13.5,
        2: 15.75,
        3: 25.5
    }.items():
        ws.row_dimensions[r].height = h

    column_widths = [
        12.57, 13.57, 12.57, 10, 10, 15.86,
        12.57, 10, 10, 8, 8, 8, 8, 8, 8, 12, 10
    ]

    for col, width in zip("ABCDEFGHIJKLMNOPQ", column_widths):
        ws.column_dimensions[col].width = width

    header_cells = [
        ("A2", "วดป", "A2:A3", thin_border()),
        ("B2", "PLรายเดือน\n=(1)+(2)+(3)", "B2:B3", thin_border()),
        ("C2", "(1) Realized", "C2:C3", thin_border()),
        ("D2", "(2) Unrealized", "D2:E2", thin_border()),
        ("F2", "(3) Commission+Vat", "F2:F3", thin_border()),
        ("G2", "4.การฝากและถอนเงิน", "G2:I2", thin_border()),
        ("J2", "จำนวนสัญญาที่เทรดในรอบ", "J2:L2", thin_border()),
        ("M2", "จำนวนสัญญาคงค้าง", "M2:O2", thin_border()),
        (
            "P2",
            "Equity ใน statement",
            "P2:P3",
            Border(
                left=thin_side,
                right=double_side,
                top=thin_side,
                bottom=thin_side
            )
        ),
        (
            "Q2",
            "Check\nต้องเท่ากับ 0",
            "Q2:Q3",
            Border(
                left=double_side,
                right=thin_side,
                top=thin_side,
                bottom=thin_side
            )
        ),
    ]

    for addr, text, merge, border in header_cells:
        c = ws[addr]
        c.value = text
        c.font = font_header
        c.alignment = align_center
        c.border = border
        ws.merge_cells(merge)

    border_cells = [
        ("E2", Border(right=thin_side, top=thin_side, bottom=thin_side)),
        ("I2", Border(right=thin_side, top=thin_side, bottom=thin_side)),
        ("L2", Border(right=thin_side, top=thin_side, bottom=thin_side)),
        ("O2", Border(right=thin_side, top=thin_side, bottom=thin_side)),
    ]

    for addr, border in border_cells:
        ws[addr].border = border

    sub_headers = [
        ("D3", "ณ สิ้นวัน"),
        ("E3", "เทียบกับ\nวันก่อนหน้า"),
        ("G3", "ฝากเงิน"),
        ("H3", "ถอนเงิน"),
        ("I3", "รวม"),
        ("J3", "Long"),
        ("K3", "Short"),
        ("L3", "Total"),
        ("M3", "Long"),
        ("N3", "Short"),
        ("O3", "Total"),
    ]

    for addr, text in sub_headers:
        c = ws[addr]
        c.value = text
        c.font = font_header
        c.alignment = align_center
        c.border = thin_border()

    if not records:
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    last_row = DATA_START_ROW + len(records) - 1

    for i, item in enumerate(records):
        r = DATA_START_ROW + i
        r1 = r - 1

        ws.row_dimensions[r].height = 15.75

        def write_cell(field, value, row_num=r):
            col = columns[field]
            cell = ws.cell(row=row_num, column=col)
            cell.value = value
            cell.font = font_for(value)
            cell.number_format = number_formats.get(col, "General")

        write_cell("date", item.get("date"))
        write_cell("realized", item.get("realized"))
        write_cell("unrealized_end_month", item.get("unrealized_end_month"))
        write_cell("commission_vat", item.get("commission_vat"))
        write_cell("deposit", item.get("deposit"))
        write_cell("withdraw", item.get("withdraw"))
        write_cell("trade_long", item.get("trade_long") or None)
        write_cell("trade_short", item.get("trade_short") or None)
        write_cell("open_long", item.get("open_long") or None)
        write_cell("open_short", item.get("open_short") or None)
        write_cell("equity", item.get("equity"))

        formulas = {
            columns["monthly_pl"]: (
                f"=C{r}+F{r}+E{r}",
                fmt_integer
            ),
            columns["net_deposit"]: (
                f"=G{r}+H{r}",
                fmt_decimal_plain
            ),
            columns["trade_total"]: (
                f"=J{r}+K{r}",
                fmt_decimal
            ),
            columns["open_total"]: (
                f"=M{r}+N{r}",
                fmt_decimal
            ),
        }

        if i > 0:
            formulas[columns["unrealized_change"]] = (
                f"=IF(F{r}<>0,D{r}-D{r1},0)",
                fmt_integer
            )
            formulas[columns["check"]] = (
                f"=IF(F{r}<>0,P{r1}+SUM(B{r})+SUM(I{r})-P{r},0)",
                fmt_integer
            )

        for col, (formula, nfmt) in formulas.items():
            cell = ws.cell(row=r, column=col)
            cell.value = formula
            cell.font = font_normal
            cell.number_format = nfmt

        for col in range(1, 18):
            ws.cell(row=r, column=col).border = Border(
                left=thin_side,
                right=thin_side
            )

        ws.cell(row=r, column=16).border = Border(right=double_side)
        ws.cell(row=r, column=17).border = Border(
            left=double_side,
            right=thin_side
        )

    total_row = last_row + 1

    ws.row_dimensions[total_row].height = 15.75

    lbl = ws.cell(row=total_row, column=1)
    lbl.value = "Total"
    lbl.font = font_bold
    lbl.fill = fill_total
    lbl.border = thin_border()

    for col in SUM_COLS:
        cl = get_column_letter(col)
        cell = ws.cell(row=total_row, column=col)

        cell.value = f"=SUM({cl}{DATA_START_ROW}:{cl}{last_row})"
        cell.font = font_bold
        cell.fill = fill_total
        cell.number_format = number_formats.get(col, "General")
        cell.border = thin_border()

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return buf


# =========================================================
# USER UI
# =========================================================

st.title("PDF Statement to Excel")
st.markdown("อัปโหลดไฟล์ PDF แล้วระบบจะสร้าง Excel ให้อัตโนมัติ")

uploaded = st.file_uploader(
    "เลือกไฟล์ PDF เลือกได้หลายไฟล์พร้อมกัน",
    type="pdf",
    accept_multiple_files=True
)


if uploaded:
    try:
        stats = load_stats()
        stats["upload_count"] += len(uploaded)
        save_stats(stats)
        log_event(f"อัปโหลด PDF {len(uploaded)} ไฟล์")
    except Exception as e:
        st.error("บันทึกข้อมูลการอัปโหลดลง Google Sheet ไม่สำเร็จ")
        st.exception(e)
        st.stop()

    st.info(f"พบ {len(uploaded)} ไฟล์ กำลังประมวลผล...")

    records = []

    for f in uploaded:
        try:
            data = read_pdf(f.read(), f.name)
            records.append(data)

            def ok(v):
                return f"{v:,.2f}" if v is not None else "—"

            st.write(
                f"{f.name} | วันที่: {data['date']} | "
                f"Realized: {ok(data['realized'])} | "
                f"Equity: {ok(data['equity'])}"
            )

        except Exception as e:
            st.error(f"อ่านไฟล์ {f.name} ไม่สำเร็จ")
            st.exception(e)

    def sort_key(item):
        parts = re.split(r"[/\-]", str(item.get("date", "")))

        try:
            if len(parts) == 3:
                d, m, y = parts
                return int(y), int(m), int(d)
        except Exception:
            pass

        return 0, 0, 0

    records.sort(key=sort_key)

    if records:
        excel_buf = make_excel(records)

        st.success(f"สร้าง Excel สำเร็จ {len(records)} แถว")

        if st.download_button(
            label="ดาวน์โหลด Excel",
            data=excel_buf,
            file_name="statement.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ):
            try:
                stats = load_stats()
                stats["download_count"] += 1
                save_stats(stats)
                log_event("ดาวน์โหลด Excel")
            except Exception as e:
                st.error("บันทึกข้อมูลการดาวน์โหลดลง Google Sheet ไม่สำเร็จ")
                st.exception(e)
    else:
        st.warning("ไม่มีไฟล์ที่ประมวลผลสำเร็จ")
